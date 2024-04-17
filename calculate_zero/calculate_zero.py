from tkinter import *
import os
import xlrd
import sys
import openpyxl

input_file_entry = None
sheet_entry = None
column_entry = None

def result_window(msg):
    window = Toplevel()
    window.geometry('300x100')
    newlabel = Label(window, text = msg)
    newlabel.pack()

    button = Button(window, text = 'Quit', command=sys.exit)
    button.pack(pady = 10)

def is_digit(data):
    try:
        int(data)
    except ValueError:
        return False

    return True

def count(worksheet, col, ws, ws_row):
    row = 1
    count_0 = 0
    try:
        while True:
            e = worksheet.cell_value(row, col)
            if (not is_digit(e)):
                with open('output.txt', 'a') as f:
                    f.write('結束. 最後還剩下 {0} 個零\n\n'.format(count_0))

                data = ['結束. 最後還剩下', str(count_0), '個零']
                for col_idx, cell_data in enumerate(data, start=1):
                    ws.cell(row=ws_row, column=col_idx, value=cell_data)
                ws_row += 2

                return ws_row

            if (e != 0):
                with open('output.txt', 'a') as f:
                    f.write('第 {0} 列是 {1}, 前面有 {2} 個零\n'.format(row + 1, int(e), count_0))

                data = ['第', str(row + 1), '列是 {0}, 前面有'.format(int(e)), str(count_0), '個零']
                for col_idx, cell_data in enumerate(data, start=1):
                    ws.cell(row=ws_row, column=col_idx, value=cell_data)
                ws_row += 1

                count_0 = 0
            else:
                count_0 += 1

            row += 1
    except IndexError as e:
        with open('output.txt', 'a') as f:
            f.write('結束. 最後還剩下 {0} 個零\n\n'.format(count_0))

        data = ['結束. 最後還剩下', str(count_0), '個零']
        for col_idx, cell_data in enumerate(data, start=1):
            ws.cell(row=ws_row, column=col_idx, value=cell_data)
        ws_row += 2

    return ws_row

def recursive_count(worksheet, start_col):
    if 'output.txt' in os.listdir():
        with open('output.txt', 'w') as f:
            pass

    if 'output.xlsx' in os.listdir():
        os.remove('output.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws_row = 1

    try:
        while True:
            val = worksheet.cell_value(0, start_col)
            if val == '':
                start_col += 1
                continue

            with open('output.txt', 'a') as f:
                try:
                    f.write('開頭第一行: {0}\n'.format(int(val)))

                    data = ['開頭第一行:', str(int(val))]
                    for col_idx, cell_data in enumerate(data, start=1):
                        ws.cell(row=ws_row, column=col_idx, value=cell_data)
                    ws_row += 1
                except ValueError as e:
                    f.write('開頭第一行: {0}\n'.format(val))

                    data = ['開頭第一行:', str(val)]
                    for col_idx, cell_data in enumerate(data, start=1):
                        ws.cell(row=ws_row, column=col_idx, value=cell_data)
                    ws_row += 1

            ws_row = count(worksheet, start_col, ws, ws_row)
            start_col += 1
    except IndexError as e:
        wb.save('output.xlsx')
        print('結束. 計算結果放在 output.txt, output.xlsx')
        result_window('結束. 計算結果放在 output.txt, output.xlsx')

def callback():
    input_file = input_file_entry.get()
    sheet = sheet_entry.get()
    column = column_entry.get()

    if not input_file:
        print('輸入檔案不能空白')
        result_window('輸入檔案不能空白')
        return

    if input_file not in os.listdir():
        print('目錄下找不到檔案: {0}'.format(input_file))
        result_window('目錄下找不到檔案: {0}'.format(input_file))
        return

    if not sheet.isnumeric():
        print('分頁不是數字: {0}'.format(sheet))
        result_window('分頁不是數字: {0}'.format(sheet))
        return

    if int(sheet) < 1:
        print('分頁數字小於1: {0}'.format(sheet))
        result_window('分頁數字小於1: {0}'.format(sheet))
        return

    if not column.isnumeric():
        print('哪一行開始不是數字: {0}'.format(column))
        result_window('哪一行開始不是數字: {0}'.format(column))
        return

    if int(column) < 1:
        print('哪一行開始數字小於1: {0}'.format(column))
        result_window('哪一行開始小於1: {0}'.format(sheet))
        return

    workbook = xlrd.open_workbook(input_file)

    worksheet = workbook.sheet_by_index(int(sheet) - 1)

    recursive_count(worksheet, int(column) - 1)

if __name__ == "__main__":
   root = Tk()
   root.geometry("300x200")

   frame = Frame(root)
   frame.pack()

   input_file_entry = Entry(frame, width = 50)
   input_file_entry.insert(0,'輸入檔案名稱, 例如: abc.xlsx')
   input_file_entry.pack(padx = 5, pady = 5)

   sheet_entry = Entry(frame, width = 50)
   sheet_entry.insert(0,'要分析哪一個分頁, 例如: 1, 2, 3, 4...')
   sheet_entry.pack(padx = 5, pady = 5)

   column_entry = Entry(frame, width = 50)
   column_entry.insert(0,'要從哪一行開始分析, 例如: 1, 10, 20...')
   column_entry.pack(padx = 5, pady = 5)

   button = Button(frame, text = "calculate", command = callback)
   button.pack()

   root.mainloop()

